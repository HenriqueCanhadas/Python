import openpyxl
import warnings
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Res Soil 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=16, max_col=16):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=17, max_col=17):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=21, max_col=21):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=25, max_col=25):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.Workbook()

# Selecione a planilha ativa (por padrão, há uma planilha chamada 'Sheet')
sheet = workbook.active
sheet.title = "Res Solo"

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    sheet.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    sheet.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    sheet.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    sheet.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    sheet.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
    sheet.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in sheet.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Res Tapwater 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=15, max_col=15):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=16, max_col=16):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=20, max_col=20):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=24, max_col=24):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.load_workbook('Tabela Epa/testearquivo.xlsx')

# Criar uma nova planilha chamada "MinhaPlanilha" no índice 2 (após a planilha padrão "Sheet")
nova_planilha = workbook.create_sheet("Res Água da torneira", 2)

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    nova_planilha.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    nova_planilha.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    nova_planilha.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    nova_planilha.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    nova_planilha.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
   nova_planilha.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in nova_planilha.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        nova_planilha.column_dimensions[column_letter].width = max(nova_planilha.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in nova_planilha.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Res Air 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=7, max_col=7):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=8, max_col=8):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=9, max_col=9):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=10, max_col=10):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.load_workbook('Tabela Epa/testearquivo.xlsx')

# Criar uma nova planilha chamada "MinhaPlanilha" no índice 2 (após a planilha padrão "Sheet")
nova_planilha = workbook.create_sheet("Res Ar", 3)

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    nova_planilha.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    nova_planilha.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    nova_planilha.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    nova_planilha.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    nova_planilha.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
   nova_planilha.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in nova_planilha.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        nova_planilha.column_dimensions[column_letter].width = max(nova_planilha.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in nova_planilha.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Soil to GW 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=15, max_col=15):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=16, max_col=16):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=20, max_col=20):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=24, max_col=24):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.load_workbook('Tabela Epa/testearquivo.xlsx')

# Criar uma nova planilha chamada "MinhaPlanilha" no índice 2 (após a planilha padrão "Sheet")
nova_planilha = workbook.create_sheet("Solo para GW 0423", 4)

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    nova_planilha.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    nova_planilha.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    nova_planilha.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    nova_planilha.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    nova_planilha.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
   nova_planilha.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in nova_planilha.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        nova_planilha.column_dimensions[column_letter].width = max(nova_planilha.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in nova_planilha.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Ind Soil 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=16, max_col=16):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=17, max_col=17):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=21, max_col=21):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=25, max_col=25):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.load_workbook('Tabela Epa/testearquivo.xlsx')

# Criar uma nova planilha chamada "MinhaPlanilha" no índice 2 (após a planilha padrão "Sheet")
nova_planilha = workbook.create_sheet("Solo Ind", 5)

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    nova_planilha.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    nova_planilha.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    nova_planilha.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    nova_planilha.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    nova_planilha.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
   nova_planilha.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in nova_planilha.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        nova_planilha.column_dimensions[column_letter].width = max(nova_planilha.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in nova_planilha.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
workbook = openpyxl.load_workbook('Tabela Epa/404086.xlsx')

sheet = workbook['Ind Air 0423']

linhaParaTerminal = ("*"*200)

nomes_quimicos = []
cas = []
cancerigeno = []
nao_cancerigeno = []

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=7, max_col=7):
    for cell in row:
        nomes_quimicos.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=8, max_col=8):
    for cell in row:
        cas.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=9, max_col=9):
    for cell in row:
        cancerigeno.append(cell.value)

# Itere pelas células da coluna G e imprima seu conteúdo
for row in sheet.iter_rows(min_row=3,min_col=10, max_col=10):
    for cell in row:
        nao_cancerigeno.append(cell.value)

workbook.close()

workbook = openpyxl.load_workbook('Tabela Epa/testearquivo.xlsx')

# Criar uma nova planilha chamada "MinhaPlanilha" no índice 2 (após a planilha padrão "Sheet")
nova_planilha = workbook.create_sheet("Ind Air", 6)

cancerigenox10 = []
valor_final = []

for i, valor in enumerate(cancerigeno):
    if isinstance(valor, (int, float)):  # Verifique se o valor é int ou float
        valor = valor * 10  # Multiplique por 10
    cancerigenox10.append(valor)
    if i == 0:
        cancerigenox10 = []
        cancerigenox10.insert(0, "Cancerigeno X 10")

for item1, item2 in zip(nao_cancerigeno, cancerigenox10):
    if isinstance(item1, (int, float)) and isinstance(item2, (int, float)):
        menor_valor = min(item1, item2)
        valor_final.append(menor_valor)
    else:
        if item2 == "Cancerigeno X 10":  # Verifica se o item1 é "Cancerigeno X 10"
            valor_final.append("Resultado")  # Substitui "Cancerigeno X 10" por "resultado"
        else:
            valor_final.append(None if item1 == "texto" or item2 == "texto" else item1 if isinstance(item1, (int, float)) else item2)

for i, valor in enumerate(nomes_quimicos, start=1):
    nova_planilha.cell(row=i, column=1, value=valor)

for i, valor in enumerate(cas, start=1):
    nova_planilha.cell(row=i, column=2, value=valor)

for i, valor in enumerate(cancerigeno, start=1):
    nova_planilha.cell(row=i, column=3, value=valor)

for i, valor in enumerate(nao_cancerigeno, start=1):
    nova_planilha.cell(row=i, column=4, value=valor)

for i, valor in enumerate(cancerigenox10, start=1):
    nova_planilha.cell(row=i, column=5, value=valor)

for i, valor in enumerate(valor_final, start=1):
   nova_planilha.cell(row=i, column=6, value=valor)

# Ajuste o tamanho das células automaticamente
for row in nova_planilha.iter_rows():
    for cell in row:
        column_letter = get_column_letter(cell.column)
        nova_planilha.column_dimensions[column_letter].width = max(nova_planilha.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in nova_planilha.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save('Tabela Epa/testearquivo.xlsx')
workbook.close
