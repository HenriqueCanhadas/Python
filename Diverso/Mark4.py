import tkinter as tk
from tkinter import filedialog
import xlwings as xw
import openpyxl
import pandas as pd
import customtkinter 

file_path = ""

def carregar_excel():
    global file_path
    global text
    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if file_path:
        global dataframe
        dataframe = pd.read_excel(file_path)
        text = customtkinter.StringVar(value="Arquivo Excel Carregado.")
        texto = customtkinter.CTkLabel(janela, textvariable=text, width=200, height=25, fg_color="blue").pack()
        print("Arquivo Excel Carregado.")
        print("-" * 100)

def rodar_programa():
    global file_path 
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl
    global valores_cliente

    if file_path == "":
        customtkinter.CTkLabel(janela,text="Nenhum Excel Carregado!").pack()
        print("Nenhum Excel Carregado!")
        print("-" * 100)
        return

    wb_xlwings = xw.Book(file_path)

    customtkinter.CTkLabel(janela,text="Rodando o Programa").pack()

    valores_cliente=[]
    valores_xlwings = []
    valores_workbook_openpyxl = []

    planilha = wb_xlwings.sheets['DADOS DO CLIENTE']

    planilha.activate()

    razao = planilha.range('B6').value
    endereco = planilha.range('B8').value
    contato= planilha.range('B10').value
    telefone = planilha.range('G10').value
    projeto = planilha.range('C14').value
    numprojeto = planilha.range('J14').value
    endeprojeto = planilha.range('C16').value
    numos= planilha.range('J18').value
    

    valores_cliente.append((razao,endereco,contato,telefone,projeto,numprojeto,endeprojeto,numos))

    print(f'Valores lidos em {planilha}:')
    print(f'Razão: {razao}')
    print(f'Endereço: {endereco}')
    print(f'Contato: {contato}')
    print(f'Telefone: {telefone}')
    print(f'Nome do Projeto: {projeto}')
    print(f'N° Projeto: {numprojeto}')
    print(f'Endereço Projeto: {endeprojeto}')
    print(f'N° OS: {numos}')
    print("-" * 100)

    for sheet_name in wb_xlwings.sheets:
        planilha = wb_xlwings.sheets[sheet_name]

        planilha.activate()

        condutividade = planilha.range('E41').value
        oxirreducao = planilha.range('G41').value
        oxigenio = planilha.range('I41').value
        ph = planilha.range('L41').value
        temperatura = planilha.range('N41').value
        turbidez = planilha.range('P41').value
        desvio1 = planilha.range('A65').value 
        desvio2 = planilha.range('A66').value 
        desvio3 = planilha.range('A67').value
        
        valores_xlwings.append((condutividade,oxirreducao,oxigenio, ph,temperatura,turbidez,desvio1,desvio2,desvio3))

        print(f'Valores lidos em {sheet_name}:')
        print(f'Condutividade: {condutividade}')
        print(f'Potencial de oxirredução: {oxirreducao}')
        print(f'Oxigênio Dissolvido: {oxigenio}')
        print(f'pH: {ph}')
        print(f'Temperatura: {temperatura}')
        print(f'Turbidez: {turbidez}')
        print(f'Desvio1: {desvio1}')
        print(f'Desvio2: {desvio2}')
        print(f'Desvio3: {desvio3}')

        print("-" * 100)

    wb_xlwings.close()

    workbook_openpyxl = openpyxl.load_workbook(file_path)

    for sheet_name in workbook_openpyxl.sheetnames:
        planilha_openpyxl = workbook_openpyxl[sheet_name]

        identificacao = planilha_openpyxl['C30'].value
        data = planilha_openpyxl['C6'].value
        horaensaio = planilha_openpyxl['H30'].value
        horaamostragem = planilha_openpyxl['K30'].value
        condicoes = planilha_openpyxl['P30'].value

        observacao = planilha_openpyxl['A69'].value
        multiparametro = planilha_openpyxl['I14'].value
        bomba = planilha_openpyxl['I16'].value
        turbidimetro = planilha_openpyxl['I18'].value
        medidor = planilha_openpyxl['I20'].value
        painel = planilha_openpyxl['I22'].value
        termometro = planilha_openpyxl['S18'].value
        vazao= planilha_openpyxl['F30'].value
        realizador = planilha_openpyxl['D8'].value

        valores_workbook_openpyxl.append((identificacao,data, horaensaio,horaamostragem,condicoes,observacao,multiparametro,bomba,turbidimetro,medidor,painel,termometro,vazao,realizador))
        
        print(f'Identificação/Aba/Guia: {identificacao}')
        print(f'Data: {data}')
        print(f'Hora do ensaio: {horaensaio}')
        print(f'Hora da amostragem: {horaamostragem}')
        print(f'Condições ambientais: {condicoes}')
        print(f'Observações: {observacao}')
        print(f'Multiparâmetro: {multiparametro}')
        print(f'Bomba: {bomba}')
        print(f'Turbidimetro: {turbidimetro}')
        print(f'Medidor de Nível: {medidor}')
        print(f'Painel controlador: {painel}')
        print(f'Termometro: {termometro}')
        print(f'Vazão: {vazao}')
        print(f'Amostragem e ensaio realizado por: {realizador}')

        print("-" * 100) 

    customtkinter.CTkLabel(janela,text="Todos os dados foram Armazenados").pack()

    print('Todos os dados foram Armazenados')

    print("-" * 100)

def salvar_excel():
    global valores_cliente
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl
    global file_path
    global status_label

    if file_path == "":
        customtkinter.CTkLabel(janela,text="Nenhum Excel Carregado!").pack()
        print("Nenhum Excel Carregado!")
        print("-" * 100)
        return
    
    if file_path:
        customtkinter.CTkLabel(janela,text="Preparando para salvar o arquivo.").pack()

        print(f"Abrindo o arquivo Excel existente na pasta:{file_path}")
        workbook_openpyxl = openpyxl.load_workbook(file_path)

        sheet = workbook_openpyxl['FINAL']

        while len(valores_cliente) > 0:
            if len(valores_cliente) > 0:
                razao,endereco,contato, telefone, projeto,numprojeto,endeprojeto,numos=valores_cliente.pop(0)
                
                sheet['B12'] = razao
                sheet['B14'] = endereco
                sheet['B16'] = contato
                sheet['G16'] = telefone
                sheet['C20'] = projeto
                sheet['J20'] = numprojeto
                sheet['C22'] = endeprojeto
                sheet['J24'] = numos

        linha = 72

        while len(valores_xlwings) > 0 and len(valores_workbook_openpyxl) > 0:
            if len(valores_xlwings) > 0:
                condutividade, oxirreducao, oxigenio, ph, temperatura, turbidez, desvio1,desvio2,desvio3 = valores_xlwings.pop(0)

                sheet[f'E{linha}'] = condutividade
                sheet[f'F{linha}'] = oxirreducao
                sheet[f'G{linha}'] = oxigenio
                sheet[f'H{linha}'] = ph
                sheet[f'I{linha}'] = temperatura
                sheet[f'J{linha}'] = turbidez

                sheet[f'L{linha}'] = desvio1
                sheet[f'M{linha}'] = desvio2
                sheet[f'N{linha}'] = desvio3
  
            if len(valores_workbook_openpyxl) > 0:
                identificacao, data, horaensaio, horaamostragem, condicoes, observacao, multiparametro, bomba, turbidimetro, medidor, painel, termometro, vazao, realizador = valores_workbook_openpyxl.pop(0)

                sheet[f'A{linha}'] = identificacao
                sheet[f'B{linha}'] = data
                sheet[f'C{linha}'] = horaensaio
                sheet[f'D{linha}'] = horaamostragem
                sheet[f'K{linha}'] = condicoes

                sheet[f'O{linha}'] = observacao
                sheet[f'P{linha}'] = multiparametro
                sheet[f'Q{linha}'] = bomba
                sheet[f'R{linha}'] = turbidimetro
                sheet[f'S{linha}'] = medidor
                sheet[f'T{linha}'] = painel
                sheet[f'U{linha}'] = termometro
                sheet[f'V{linha}'] = vazao
                sheet[f'W{linha}'] = realizador
            
            linha += 1

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if save_path:
        workbook_openpyxl.save(save_path)
        customtkinter.CTkLabel(janela,text="Excel salvo com sucesso!").pack()

        print("Excel salvo com sucesso!")
        print("-" * 100)

janela=customtkinter.CTk()
janela.geometry("250x250")
janela.title("SERVMAR")

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

load_button = customtkinter.CTkButton(janela, text="Carregar Excel", command=carregar_excel)
load_button.pack(pady=20)

run_button = customtkinter.CTkButton(janela, text="Rodar Programa", command=rodar_programa)
run_button.pack(pady=20)

save_button = customtkinter.CTkButton(janela, text="Salvar Excel", command=salvar_excel)
save_button.pack(pady=20)

text = customtkinter.StringVar(value="")
texto = customtkinter.CTkLabel(janela, textvariable=text, width=200, height=25, fg_color="blue").pack()

janela.mainloop()