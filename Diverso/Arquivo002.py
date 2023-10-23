import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Criar um DataFrame de exemplo
data = {
    'Nome': ['Alice', 'Bob', 'Charlie'],
    'Idade': [25, 30, 35],
    'Cidade': ['Nova York', 'Los Angeles', 'Chicago']
}
df = pd.DataFrame(data)

# Criar um novo arquivo Excel com o openpyxl
wb = Workbook()
ws = wb.active

# Escrever os dados do DataFrame no arquivo Excel
for r_idx, row in enumerate(df.itertuples(), start=1):
    for c_idx, value in enumerate(row[1:], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Mesclar células na primeira linha
ws.merge_cells('A1:C1')

# Formatar a célula mesclada (nesse caso, apenas centralizar o texto)
cell = ws['A1']
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.font = Font(bold=True)

# Para definir o texto em vermelho
cell = ws['B2']  # Substitua 'B1' pela célula desejada
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.font = Font(color='FF0000')  # 'FF0000' é a cor vermelha em formato hexadecimal

# ...

# Salvar o arquivo Excel
wb.save('tabela_excel_formatada.xlsx')
