import openpyxl
from copy import copy


# Abra o arquivo Excel existente
workbook = openpyxl.load_workbook('Tabela Servmar/teste.xlsx')

# Especifique a planilha em que você deseja trabalhar (pode ser pelo nome ou índice)
sheet = workbook['TESTEGUIA']

# Defina a quantidade de colunas a serem inseridas
num_colunas_inserir = 10  # Altere o número conforme necessário

# Mesclar as células nas linhas 5 e 6
sheet.merge_cells('E5:F5')
sheet.merge_cells('E6:F6')

# Loop para inserir várias colunas
for _ in range(num_colunas_inserir):
    sheet.insert_cols(6)  # Inserir uma nova coluna entre E e F (coluna 6)

    # Copiar os valores e formatações da coluna E para a nova coluna
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            new_cell = sheet.cell(row=cell.row, column=6, value=cell.value)
            new_cell._style = copy(cell._style)  # Copiar formatações

# Ajustar a largura das colunas a partir da coluna E em diante
for column in sheet.iter_cols(min_col=5):
    max_length = 0
    for cell in column:
        if cell.coordinate in sheet.merged_cells:
            continue
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    column_letter = column[0].column_letter
    sheet.column_dimensions[column_letter].width = adjusted_width


# Salve o arquivo Excel modificado
workbook.save('Tabela Servmar/teste1.xlsx')