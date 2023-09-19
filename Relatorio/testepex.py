import xlwings as xw
import openpyxl
"""""
# Abrir o arquivo Excel com xlwings
wb = xw.Book('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

# Acessar uma planilha específica
planilha = wb.sheets['PMN-03']

valor_celula = planilha.range('E41').value
print(f'Valor da célula A1: {valor_celula}')

wb.close()

# Agora, você pode usar openpyxl para operações adicionais, como leitura ou escrita de dados
# Exemplo:
wb_openpyxl = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')
planilha_openpyxl = wb_openpyxl['PMN-03']

# Ler ou escrever dados com openpyxl
valor_celula_openpyxl = planilha_openpyxl['E41'].value
print(f'Valor da célula B1 (usando openpyxl): {valor_celula_openpyxl}')
"""
#----------------------------------------------------------------------------------------------------------

import xlwings as xw
import openpyxl

    # Abrir o arquivo Excel com xlwings
wb_xlwings = xw.Book('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

nomes_das_guias = [sheet.name for sheet in wb_xlwings.sheets]
for sheet_name in nomes_das_guias:
    planilha = wb_xlwings.sheets[sheet_name]

    # Faça algo com a planilha, por exemplo, leia o valor de uma célula
    valor_celula = planilha.range('A1').value

    # Imprima o valor da célula e o nome da planilha
    print(f'Valor da célula A1 na planilha {sheet_name}: {valor_celula}')

    # Acessar a planilha desejada com xlwings
planilha_xlwings = wb_xlwings.sheets['PMN-03']

    # Ler o valor da célula com xlwings
valor_celula_xlwings = planilha_xlwings.range('E41').value

    # Fechar o arquivo Excel com xlwings
wb_xlwings.close()

    # Abrir o arquivo Excel com openpyxl
wb_openpyxl = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    # Acessar a planilha desejada com openpyxl
planilha_openpyxl = wb_openpyxl['PMN-03']

    # Escrever o valor lido de xlwings em uma célula correspondente com openpyxl
planilha_openpyxl['E41'].value = valor_celula_xlwings

    # Salvar as alterações com openpyxl
wb_openpyxl.save('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    # Fechar o arquivo Excel com openpyxl
wb_openpyxl.close()

    # Imprimir o valor lido em openpyxl
print(f'Valor da célula E41 em openpyxl: {planilha_openpyxl["E41"].value}')


#----------------------------------------------------------------------------------------------------------
"""""
import xlwings as xw
import openpyxl

wb = xw.Book('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

for sheet_name in wb.sheets:

    wb = xw.Book('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    planilha = wb.sheets[sheet_name] 

    condutividade = planilha.range('E41').value
    oxirreducao = planilha.range('G41').value
    oxigenio = planilha.range('I41').value
    ph = planilha.range('L41').value
    temperatura = planilha.range('N41').value
    turbidez = planilha.range('P41').value
   
    wb.close()

    workbook = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    # Acessar a planilha desejada com openpyxl
    planilha = wb[sheet_name]

    # Escrever o valor lido de xlwings em uma célula correspondente com openpyxl
    planilha['E41'].value = condutividade
    planilha['G41'].value = oxirreducao
    planilha['I41'].value = oxigenio
    planilha['L41'].value = ph
    planilha['N41'].value = temperatura
    planilha['P41'].value = turbidez

"""