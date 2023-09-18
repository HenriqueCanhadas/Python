import openpyxl
import pandas as pd

workbook = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

n=2

for sheet_name in workbook.sheetnames:

    workbook = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    tabela = pd.read_excel('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx', sheet_name=sheet_name)

    sheet = workbook[sheet_name]

    multiparametro=sheet['I14'].value
    turbidimetro =sheet['I18'].value
    identificacao  =sheet['C30'].value
    data =sheet['C6'].value
    horaensaio =sheet['H30'].value
    horaamostragem = sheet['K30'].value
    condutividade = tabela.iat[39, 4]#E41
    oxirreducao = tabela.iat[39, 6]#G41
    oxigenio = tabela.iat[39, 8]#I41
    ph = tabela.iat[39, 11]#L41
    temperatura = tabela.iat[39, 13]#N41
    turbidez = tabela.iat[39, 15]#P41
    condicoes =sheet['P30'].value

    workbook = openpyxl.load_workbook('Relatorio\Testeopen.xlsx')

    sheet = workbook['aba1']

    sheet[f'A{n}']= multiparametro
    sheet[f'B{n}']=turbidimetro
    sheet[f'C{n}']=identificacao
    sheet[f'D{n}']=data
    sheet[f'E{n}']= horaensaio
    sheet[f'F{n}']=horaamostragem
    sheet[f'G{n}']=condutividade
    sheet[f'H{n}']=oxirreducao
    sheet[f'I{n}']=oxigenio
    sheet[f'J{n}']=ph
    sheet[f'K{n}']=temperatura
    sheet[f'L{n}']=turbidez
    sheet[f'M{n}']=condicoes

    n+=1

    print(f'N esta em:{n}')

    workbook.save('Relatorio\Testeopen.xlsx')

    print(multiparametro)
    print(turbidimetro)
    print(identificacao)
    print(data)
    print(horaensaio)
    print(horaamostragem)
    print(condutividade)
    print(oxirreducao)
    print(oxigenio)
    print(ph)
    print(temperatura)
    print(turbidez)
    print(condicoes)




