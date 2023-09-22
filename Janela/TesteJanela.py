import tkinter as tk
from tkinter import filedialog
import xlwings as xw
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import deepcopy

# Variável global para armazenar o caminho do arquivo Excel
file_path = ""

# Função para carregar o arquivo Excel
def carregar_excel():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
    if file_path:
        global dataframe
        dataframe = pd.read_excel(file_path)
        status_label.config(text="Arquivo Excel carregado com sucesso!")


# Função para rodar o programa
def rodar_programa():
    global file_path  # Acessa a variável global file_path
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl

    wb_xlwings = xw.Book(file_path)

    valores_xlwings = []
    valores_workbook_openpyxl = []

    for sheet_name in wb_xlwings.sheets:

        planilha = wb_xlwings.sheets[sheet_name]

        planilha.activate()

        condutividade = planilha.range('E41').value
        oxirreducao = planilha.range('G41').value
        oxigenio = planilha.range('I41').value
        ph = planilha.range('L41').value
        temperatura = planilha.range('N41').value
        turbidez = planilha.range('P41').value

        valores_xlwings.append((condutividade, oxirreducao, oxigenio, ph, temperatura, turbidez))

        print(f'Valores lidos em {sheet_name}:')
        print(f'Condutividade: {condutividade}')
        print(f'Potencial de oxirredução: {oxirreducao}')
        print(f'Oxigênio Dissolvido: {oxigenio}')
        print(f'pH: {ph}')
        print(f'Temperatura: {temperatura}')
        print(f'Turbidez: {turbidez}')
        print("-" * 100)

    wb_xlwings.close()

    workbook_openpyxl = openpyxl.load_workbook(file_path)

    for sheet_name in workbook_openpyxl.sheetnames:
        planilha_openpyxl = workbook_openpyxl[sheet_name]

        identificacao = planilha_openpyxl['C30'].value
        data = planilha_openpyxl['C6'].value
        horaensaio = planilha_openpyxl['H30'].value
        horaamostragem = planilha_openpyxl['K30'].value
        condicoes = planilha_openpyxl['P30'].value

        valores_workbook_openpyxl.append((identificacao, data, horaensaio, horaamostragem, condicoes))


        print(f'Identificação/Aba/Guia: {identificacao}')
        print(f'Data: {data}')
        print(f'Hora do ensaio: {horaensaio}')
        print(f'Hora da amostragem: {horaamostragem}')
        print(f'Condições ambientais: {condicoes}')

        print("-" * 100)

    status_label.config(text="Todos os dados foram Armazenados")

    print('Todos os dados foram Armazenados')

    print("-" * 100)



def salvar_excel():
    global valores_xlwings
    global valores_workbook_openpyxl
    global workbook_openpyxl
    global file_path
    global status_label

    if file_path:
        # Abrir o arquivo Excel original
        print(f"Abrindo o arquivo Excel existente em: {file_path}")
        workbook_original = openpyxl.load_workbook(file_path)

        # Criar uma cópia exata do arquivo Excel original
        workbook_copia = deepcopy(workbook_original)

        # Acesse a planilha 'FINAL' na cópia (ou ajuste conforme necessário)
        sheet = workbook_copia.active
        sheet.title = 'FINAL'

        linha = 69

        while len(valores_xlwings) > 0 and len(valores_workbook_openpyxl) > 0:
            if len(valores_xlwings) > 0:
                condutividade, oxirreducao, oxigenio, ph, temperatura, turbidez = valores_xlwings.pop(0)

                sheet[f'E{linha}'] = condutividade
                sheet[f'F{linha}'] = oxirreducao
                sheet[f'G{linha}'] = oxigenio
                sheet[f'H{linha}'] = ph
                sheet[f'I{linha}'] = temperatura
                sheet[f'J{linha}'] = turbidez
            
            if len(valores_workbook_openpyxl) > 0:
                identificacao, data, horaensaio, horaamostragem, condicoes = valores_workbook_openpyxl.pop(0)

                sheet[f'A{linha}'] = identificacao
                sheet[f'B{linha}'] = data
                sheet[f'C{linha}'] = horaensaio
                sheet[f'D{linha}'] = horaamostragem
                sheet[f'K{linha}'] = condicoes
            
            linha += 1

        # Salvar o arquivo com um novo nome
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
        if save_path:
            workbook_copia.save(save_path)
            status_label.config(text="Novo arquivo Excel salvo com sucesso!")

        print("Processo concluído com sucesso.")


# Criação da janela principal
root = tk.Tk()
root.title("Nome Do Programa")

# Definir tamanho mínimo para a janela (width x height)
root.minsize(250, 250)  # Ajuste os valores de largura e altura conforme necessário

# Botão para carregar o Excel
load_button = tk.Button(root, text="Carregar Excel", command=carregar_excel)
load_button.pack(pady=20)

# Botão para rodar o programa
load_button = tk.Button(root, text="Rodar Programa", command=rodar_programa)
load_button.pack(pady=20)

# Botão para salvar o Excel
save_button = tk.Button(root, text="Salvar Excel", command=salvar_excel)
save_button.pack(pady=20)

# Rótulo de status
status_label = tk.Label(root, text="")
status_label.pack()

root.mainloop()