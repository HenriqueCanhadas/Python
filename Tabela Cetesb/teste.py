import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook('Tabela Cetesb/arquivoteste.xlsx')

sheet = workbook['Tabela']

linhaParaTerminal = ("*"*200)

nomes_parametros = []

nomes_quimicos = []  

coordenada = []

cas = []
vl_agricola = []
vl_residencial = []
vl_industrial = []
vl_aguasub = []

nomes = [
    "INORGÂNICOS",
    "HIDROCARBONETOS AROMÁTICOS VOLÁTEIS",
    "HIDROCARBONETOS POLICÍCLICOS AROMÁTICOS",
    "BENZENOS CLORADOS",
    "ETANOS CLORADOS",
    "ETENOS CLORADOS",
    "METANOS CLORADOS",
    "FENÓIS CLORADOS",
    "FENÓIS NÃO CLORADOS",
    "ÉSTERES FTÁLICOS",
    "PESTICIDAS",
    "OUTROS",
    "VRQ – Valor de Referência de Qualidade; VP – Valor de Prevenção; VI – Valor de Intervenção",
]

for row in sheet.iter_rows():
    for cell in row:
        if cell.value is not None:
            for nome in nomes:
                if nome.lower() in str(cell.value).lower():
                    nomes_parametros.append(cell.value)
                    coordenada.append(cell.coordinate)

first_coord = coordenada[0]
last_coord = coordenada[-1]

start_cell = sheet[first_coord]
end_cell = sheet[last_coord]

def obter_valores():

    for row in sheet.iter_rows(min_row=start_cell.row, max_row=end_cell.row, min_col=start_cell.column, max_col=end_cell.column):
        for cell in row:
                if cell.value is not None:
                    if "VRQ – Valor de Referência de Qualidade; VP – Valor de Prevenção; VI – Valor de Intervenção" in str(cell.value):
                        break  
                nomes_quimicos.append(cell.value)

    #Laço para pegar os valores da celula B,C,D
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=2, max_col=4):  #min_col=2 == B, max_col=4 == D
        for cell in row:
            # Verifique se o valor da coluna B é None
            if cell.column == 2:  # Coluna B
                if cell.value is None:
                    # Encontre o próximo valor nas colunas C (3) e D (4) na mesma linha
                    valor_coluna_C = None
                    valor_coluna_D = None
                    for i in range(1, len(row)):
                        if row[i].column == 3:  # Coluna C
                            valor_coluna_C  = row[i].value
                        elif row[i].column == 4:  # Coluna D
                            valor_coluna_D = row[i].value

                    # Escolha entre o valor da coluna C ou D
                    if valor_coluna_C is not None:
                        cas.append(valor_coluna_C)
                    elif valor_coluna_D is not None:
                        cas.append(valor_coluna_D)
                else:
                    cas.append(cell.value)

    #Laço para pegar os valores da celula H,I,J
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=8, max_col=10):   #min_col=8 == H, max_col=10 == J
        for cell in row:
            if cell.column == 8:  
                if cell.value is None:
                    valor_coluna_I = None
                    valor_coluna_J = None
                    for i in range(1, len(row)):
                        if row[i].column == 9:  
                            valor_coluna_I = row[i].value
                        elif row[i].column == 10:  
                            valor_coluna_J = row[i].value
                    if valor_coluna_I is not None:
                        vl_agricola.append(valor_coluna_I)
                    elif valor_coluna_J is not None:
                        vl_agricola.append(valor_coluna_J)
                else:
                    vl_agricola.append(cell.value)

    #Laço para pegar os valores da celula J,k,L
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=10, max_col=12):  #min_col=10 == J, max_col=12 == L
        for cell in row:
            if cell.column == 10:
                if cell.value is None:
                    valor_coluna_I = None
                    valor_coluna_J = None
                    for i in range(1, len(row)):
                        if row[i].column == 11: 
                            valor_coluna_I = row[i].value
                        elif row[i].column == 12:  
                            valor_coluna_J = row[i].value
                    if valor_coluna_I is not None:
                        vl_residencial.append(valor_coluna_I)
                    elif valor_coluna_J is not None:
                        vl_residencial.append(valor_coluna_J)
                else:
                    vl_residencial.append(cell.value)

    #Laço para pegar os valores da celula L,M,N
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=12, max_col=14):  #min_col=12 == L, max_col=14 == N
        for cell in row:
            if cell.column == 12:
                if cell.value is None:
                    valor_coluna_I = None
                    valor_coluna_J = None
                    for i in range(1, len(row)):
                        if row[i].column == 13:  
                            valor_coluna_I = row[i].value
                        elif row[i].column == 14:
                            valor_coluna_J = row[i].value
                    if valor_coluna_I is not None:
                        vl_industrial.append(valor_coluna_I)
                    elif valor_coluna_J is not None:
                        vl_industrial.append(valor_coluna_J)
                else:
                    vl_industrial.append(cell.value)

    #Laço para pegar os valores da celula O,P,Q
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=15, max_col=17):  #min_col=15 == O, max_col=17 == Q
        for cell in row:
            # Verifique se o valor da coluna H é None
            if cell.column == 15:  # Coluna H
                if cell.value is None:
                    # Encontre o próximo valor nas colunas I (9) e J (10) na mesma linha
                    valor_coluna_I = None
                    valor_coluna_J = None
                    for i in range(1, len(row)):
                        if row[i].column == 16:  # Coluna I
                            valor_coluna_I = row[i].value
                        elif row[i].column == 17:  # Coluna J
                            valor_coluna_J = row[i].value

                    # Escolha entre o valor da coluna I ou J
                    if valor_coluna_I is not None:
                        vl_aguasub.append(valor_coluna_I)
                    elif valor_coluna_J is not None:
                        vl_aguasub.append(valor_coluna_J)
                else:
                    vl_aguasub.append(cell.value)

    for row in sheet.iter_rows(min_row=82, max_row=sheet.max_row, min_col=14, max_col=16):  #min_col=15 == O, max_col=17 == Q
        for cell in row:
            # Verifique se o valor da coluna H é None
            if cell.column == 14:  # Coluna H
                if cell.value is None:
                    # Encontre o próximo valor nas colunas I (9) e J (10) na mesma linha
                    valor_coluna_I = None
                    valor_coluna_J = None
                    for i in range(1, len(row)):
                        if row[i].column == 15:  # Coluna I
                            valor_coluna_I = row[i].value
                        elif row[i].column == 16:  # Coluna J
                            valor_coluna_J = row[i].value

                    # Escolha entre o valor da coluna I ou J
                    if valor_coluna_I is not None:
                        vl_aguasub.append(valor_coluna_I)
                    elif valor_coluna_J is not None:
                        vl_aguasub.append(valor_coluna_J)
                else:
                    vl_aguasub.append(cell.value)

obter_valores()

print(vl_aguasub)

quantidade = len(vl_aguasub)

print(quantidade)

#Fecha o Arquivo teste
workbook.close()

# Crie um novo arquivo Excel
workbook = openpyxl.Workbook()

# Selecione a planilha ativa (por padrão, há uma planilha chamada 'Sheet')
sheet = workbook.active

fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Adicionar os valores à planilha
for i, valor in enumerate(nomes_quimicos):
    sheet.cell(row=i + 1, column=1, value=valor)

    # Verifique se o valor atual está na lista de nomes para mesclar células
    if valor in nomes:
        # Mesclar as células da coluna A até a F
        sheet.merge_cells(start_row=i + 1, start_column=1, end_row=i + 1, end_column=6)
        # Aplicar o estilo de preenchimento à primeira célula do intervalo mesclado
        sheet.cell(row=i + 1, column=1).fill = fill

linha_atual = 1

# Loop para a coluna B (cas)
while cas:
    valor = cas.pop(0)
    while sheet.cell(row=linha_atual, column=2).coordinate in sheet.merged_cells:
        linha_atual += 1

    # Insira o valor na célula da coluna B e vá para a próxima linha
    sheet.cell(row=linha_atual, column=2, value=valor)
    linha_atual += 1
    print(f"Inserindo '{valor}' na célula B{linha_atual - 1}")

# Reinicie a linha atual para a coluna C (vl_agricola)
linha_atual = 1

# Loop para a coluna C (vl_agricola)
while vl_agricola:
    valor = vl_agricola.pop(0)
    while sheet.cell(row=linha_atual, column=3).coordinate in sheet.merged_cells:
        linha_atual += 1

    # Insira o valor na célula da coluna C e vá para a próxima linha
    sheet.cell(row=linha_atual, column=3, value=valor)
    linha_atual += 1
    print(f"Inserindo '{valor}' na célula C{linha_atual - 1}")

linha_atual = 1

# Loop para a coluna D (vl_residencial)
while vl_residencial:
    valor = vl_residencial.pop(0)
    while sheet.cell(row=linha_atual, column=4).coordinate in sheet.merged_cells:
        linha_atual += 1

    # Insira o valor na célula da coluna D e vá para a próxima linha
    sheet.cell(row=linha_atual, column=4, value=valor)
    linha_atual += 1
    print(f"Inserindo '{valor}' na célula B{linha_atual - 1}")

# Reinicie a linha atual para a coluna C (vl_residencial)
linha_atual = 1

# Loop para a coluna E (vl_industrial)
while vl_industrial:
    valor = vl_industrial.pop(0)
    while sheet.cell(row=linha_atual, column=5).coordinate in sheet.merged_cells:
        linha_atual += 1

    # Insira o valor na célula da coluna C e vá para a próxima linha
    sheet.cell(row=linha_atual, column=5, value=valor)
    linha_atual += 1
    print(f"Inserindo '{valor}' na célula C{linha_atual - 1}")

linha_atual = 1

# Loop para a coluna E (vl_industrial)
while vl_aguasub:
    valor = vl_aguasub.pop(0)
    while sheet.cell(row=linha_atual, column=6).coordinate in sheet.merged_cells:
        linha_atual += 1
    # Verifique se o valor é igual ao valor específico
    if valor == "20 (b)":
        for _ in range(2):
            sheet.cell(row=linha_atual, column=6, value=valor)
            print(f"Inserindo '{valor}' 3 vezes na célula B{linha_atual}")
            linha_atual += 1
    
    if valor == "50 (b)":
        for _ in range(1):
            sheet.cell(row=linha_atual, column=6, value=valor)
            print(f"Inserindo '{valor}' 3 vezes na célula B{linha_atual}")
            linha_atual += 1

    if valor == "0,03 (b)":
        for _ in range(1):
            sheet.cell(row=linha_atual, column=6, value=valor)
            print(f"Inserindo '{valor}' 3 vezes na célula B{linha_atual}")
            linha_atual += 1

    if valor == "1 (b)":
        for _ in range(3):
            sheet.cell(row=linha_atual, column=6, value=valor)
            print(f"Inserindo '{valor}' 3 vezes na célula B{linha_atual}")
            linha_atual += 1
    
    else:
        # Insira o valor na célula da coluna B e vá para a próxima linha
        sheet.cell(row=linha_atual, column=6, value=valor)
        print(f"Inserindo '{valor}' na célula B{linha_atual}")
        linha_atual += 1


# Ajuste o tamanho das células automaticamente
for row in sheet.iter_rows():
    for cell in row:
        # Verifica se a célula está mesclada
        if cell.coordinate in sheet.merged_cells:
            continue
        column_letter = get_column_letter(cell.column)
        sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width, len(str(cell.value)) + 2)

# Alinhe as células centralmente
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
# Salve o arquivo Excel
workbook.save('Tabela Cetesb/meuarquivo.xlsx')
