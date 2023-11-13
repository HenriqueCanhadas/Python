import pandas as pd
import openpyxl
from openpyxl.styles import Font

# Abre o arquivo excel que possue os resultados das analises que o laboratório enviou
arquivo_excel_lab = openpyxl.load_workbook(
    r'C:\Users\vitor.lucas\Desktop\Teste Analise\Export_Toyota_SBC_Geral2.xlsx')  # Antigo workbook

planilha = arquivo_excel_lab['Amostras Resultados']  # Antiga sheet

pm = []  # antigo id
tipo_de_amostra = []  # antigo tipo
data_hora = []  # antiga data
todos = []  # antigo id_parametros


# Função que assume a primeira linha como cabeçalho, remove os espaços das celulas,  e adiciona nas respectivas listas pm, tipo_de_amostra, data_hora e também, cria uma lista com todas as listas juntas 
def coletar_dados():
    global arquivo_excel_lab
    # Assume que a primeira linha é a linha de cabeçalho, verifica se as celulas possuem valor de texto e remove os espaços em branco
    # Para cada linha da planilha 'Amostras Resultados' no intervalo de linha 2 a coluna 4, cada celula na linha, remove os espaços dos valores
    for linha in planilha.iter_rows(min_row=2, min_col=4, max_col=4):
        for celula in linha:
            if celula.value is not None and isinstance(celula.value, str):
                celula.value = celula.value.replace(" ", "")

    # Adiciona os valores às listas quando a próxima célula for nula ou diferente
    # O "+1" serve para incluir a última linha
    # Para cada linha, armazena a celula da coluna
    for linha in range(2, planilha.max_row + 1):
        celula_pm = planilha.cell(row=linha, column=1).value  # Antiga cell1
        celula_seguinte = planilha.cell(row=linha+1, column=1).value  # Antiga cell2

        if celula_pm is not None and (celula_seguinte is None or celula_pm != celula_seguinte):
            pm.append(planilha.cell(row=linha, column=1).value)
            tipo_de_amostra.append(planilha.cell(row=linha, column=2).value)
            data_hora.append(planilha.cell(row=linha, column=3).value)

    # Adiciona novas variáveis para criar a lista com todos os dados juntos
    for linha in range(2, planilha.max_row + 1):
        # Valor da coluna 'A' IDENTIFICAÇÃO DA AMOSTRA
        celula_pm = planilha.cell(row=linha, column=1).value
        # Valor da coluna 'D' ANALISE
        celula_analise = planilha.cell(row=linha, column=4).value
        # Valor da coluna 'E' RESULTADO
        celula_resultado = planilha.cell(row=linha, column=5).value
        # Valor da coluna 'F' UNIDADE
        celula_unidade = planilha.cell(row=linha, column=6).value

        if celula_pm is not None and (celula_analise is not None or celula_resultado is not None or celula_unidade is not None):
            todos.append([celula_pm, [celula_analise, celula_resultado, celula_unidade]])

    print("Leu o Coletar Dados", end="\n")
    print(pm)

def criar_excel():
    global arquivo_excel_lab
    global arquivo_excel_novo
    global todos

    # Cria um DataFrame vazio
    data_frame = pd.DataFrame()  # Antigo df

    quantidade_pm = max(len(pm), len(tipo_de_amostra))

    data_frame = pd.DataFrame({'cont': range(quantidade_pm), 'Identificação da Amostra': pm,
                              'Tipo de amostra': tipo_de_amostra, 'Data e Hora': data_hora})

    # Salva o DataFrame em um arquivo Excel
    arquivo_excel_novo = r'C:\Users\vitor.lucas\Desktop\Teste Analise\TesteExcel.xlsx'
    data_frame.to_excel(arquivo_excel_novo, index=False)

    for nome_da_aba in pm:
        nova_aba = arquivo_excel_lab.create_sheet(title=nome_da_aba)
        linha_start = 1  # Antigo linha

    # Salva o DataFrame em um arquivo Excel
    arquivo_excel_novo = r'C:\Users\vitor.lucas\Desktop\Teste Analise\TesteExcel.xlsx'
    data_frame.to_excel(arquivo_excel_novo, index=False)

    arquivo_excel_lab = openpyxl.load_workbook(arquivo_excel_novo)

    for nome_da_aba in pm:
        nova_aba = arquivo_excel_lab.create_sheet(title=nome_da_aba)
        linha_start = 1  # Antigo linha

        for item_de_todos in todos:
            if item_de_todos[0] == nome_da_aba:
                for indice, subitem_de_todos in enumerate(item_de_todos[1]):
                    nova_aba.cell(row=linha_start, column=indice +
                                  1, value=subitem_de_todos)
                linha_start += 1

    arquivo_excel_lab.save(arquivo_excel_novo)

    print("Leu o Criar Excel", end="\n")


def alinhar_celula():
    global arquivo_excel_lab
    global arquivo_excel_novo

    arquivo_excel_lab = openpyxl.load_workbook(arquivo_excel_novo)

    # Ajusta o alinhamento horizontal
    for planilha in arquivo_excel_lab:
        for linha in planilha.iter_rows(min_row=1, max_row=planilha.max_row, min_col=1, max_col=planilha.max_column):
            for alinhar_celula in linha:
                alinhar_celula.alignment = openpyxl.styles.Alignment(
                    horizontal='center')

        # Ajusta o alinhamento vertical
        for linha in planilha.iter_rows(min_row=1, max_row=planilha.max_row, min_col=1, max_col=planilha.max_column):
            for alinhar_celula in linha:
                alinhar_celula.alignment = openpyxl.styles.Alignment(
                    vertical='center')

        # Ajusta a largura das colunas de acordo com o conteúdo
        for coluna in planilha.columns:
            tamanho_maximo = 0
            coluna = [celula for celula in coluna]
            for celula in coluna:
                try:
                    if len(str(celula.value)) > tamanho_maximo:
                        tamanho_maximo = len(celula.value)
                except:
                    pass
            largura_ajustada = (tamanho_maximo + 2)
            planilha.column_dimensions[coluna[0].column_letter].width = largura_ajustada

    # Salva as alterações
    arquivo_excel_lab.save(arquivo_excel_novo)
    # Fecha o arquivo
    arquivo_excel_lab.close()

    print("Leu o Alinhar Celulas", end="\n")


def verificar_valor():
    global arquivo_excel_lab
    global arquivo_excel_novo

    # Define a cor da fonte
    fonte_cinza = Font(color="FF808080")
    fonte_vermelha = Font(color="FF0000")

    for planilha in arquivo_excel_lab.sheetnames[1:]:
        planilha = arquivo_excel_lab[planilha]

        coluna_b = 'B'  # Mostra a coluna que vai ser alterada é a 'B'
        for celula in planilha[coluna_b]: # Antiga column_letter
            if celula.value and "<" in celula.value:
                celula.font = fonte_cinza
            else:
                celula.font = fonte_vermelha

    # Salva as alterações
    arquivo_excel_lab.save(
        r'C:\Users\vitor.lucas\Desktop\Teste Analise\TesteExcel.xlsx')

    print("Leu o Verificar Valor", end="\n")


coletar_dados()

criar_excel()

alinhar_celula()

verificar_valor()
