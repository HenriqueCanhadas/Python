import pandas as pd
import openpyxl
import xlwings

workbook = openpyxl.load_workbook('TestePivot/Export_Toyota_SBC_Geral2.xlsx')

sheet = workbook['Amostras Resultados']

id=[]
tipo=[]
data=[]

column_number = 1  # Número da coluna desejada (1 para A, 2 para B, etc.)

for row in range(2, sheet.max_row + 1):  # Note o "+ 1" para incluir a última linha.
    cell1 = sheet.cell(row=row, column=column_number).value
    cell2 = sheet.cell(row=row + 1, column=column_number).value

    if cell1 is not None and (cell2 is None or cell1 != cell2):
        # Adicionar os valores às listas quando a próxima célula for nula ou diferente
        id.append(sheet.cell(row=row, column=1).value)
        tipo.append(sheet.cell(row=row, column=2).value)
        data.append(sheet.cell(row=row, column=3).value)

id_parametros = []

for row in range(2, sheet.max_row + 1):  # Note o "+ 1" para incluir a última linha.
    cell1 = sheet.cell(row=row, column=column_number).value
    cellD = sheet.cell(row=row, column=4).value  # Valor da coluna 'D'
    cellE = sheet.cell(row=row, column=5).value  # Valor da coluna 'E'
    cellF = sheet.cell(row=row, column=6).value  # Valor da coluna 'F'

    if cell1 is not None and (cellD is not None or cellE is not None or cellF is not None):
        # Adicionar os valores à lista id no formato desejado
        id_parametros.append([cell1, [cellD, cellE, cellF]])

print(id_parametros[0])

# Criar um DataFrame vazio
df = pd.DataFrame()

tamanho_lista = max(len(id), len(tipo))

df = pd.DataFrame({'cont': range(tamanho_lista), 'Id Amostra': id, 'Tipo': tipo})

# Salvar o DataFrame em um arquivo Excel
arquivo_excel = 'TestePivot/TesteExcel.xlsx'
df.to_excel(arquivo_excel, index=False)

wb = openpyxl.load_workbook(arquivo_excel)

for nome_da_aba in id:
    nova_aba = workbook.create_sheet(title=nome_da_aba)
    linha = 1
    

# Salvar o DataFrame em um arquivo Excel
arquivo_excel = 'TestePivot/TesteExcel17.xlsx'
df.to_excel(arquivo_excel, index=False)


workbook = openpyxl.load_workbook(arquivo_excel)

for nome_da_aba in id:
    nova_aba = workbook.create_sheet(title=nome_da_aba)
    linha = 1
    
    # Itere sobre os dados e insira-os nas células da nova aba nas colunas A, B e C
    for item in id_parametros:
        if item[0] == nome_da_aba:
            for i, subitem in enumerate(item[1]):
                nova_aba.cell(row=linha, column=i+1, value=subitem)
            linha+=1

workbook.save(arquivo_excel)

print("Arquivo Excel criado comd sucesso.")

def alinha_celula(arquivo_excel):

    workbook = openpyxl.load_workbook(arquivo_excel)

    for sheet in workbook:
        # 1. Defina o alinhamento horizontal (por exemplo, 'center')
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # 2. Defina o alinhamento vertical (por exemplo, 'middle')
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(vertical='center')

        # 3. Ajuste automaticamente a largura das colunas para que o conteúdo se ajuste
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Salve as alterações
    workbook.save(arquivo_excel)

    # Feche o arquivo
    workbook.close()


alinha_celula(arquivo_excel)

print(f'Valores lidos em {sheet}:')

