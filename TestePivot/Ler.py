import pandas
import openpyxl
import xlwings

workbook = openpyxl.load_workbook('TestePivot/Export_Toyota_SBC_Geral.xlsx')

sheet = workbook['Amostras Resultados (2)']

''' 
    def printLinhaLaboratorio(linha):
        for linha in sheet.iter_rows(min_row=linha, max_row=linha):
            for cell in linha:
                print(cell.column,cell.value, end=', ')
            print()
 
    contador = 0


    for row in sheet.iter_rows(min_col=5, max_col=5, max_row=sheet.max_row):

        for cell in row:
            if cell.value is not None and "<" not in str(cell.value):
                linha = cell.row
                printLinhaLaboratorio(linha)
                contador += 1

    print("Total:",contador)
'''

id=[]
tipo=[]
data=[]



column_number = 1  # Número da coluna desejada (1 para A, 2 para B, etc.)

for row in range(1, sheet.max_row + 1):  # Note o "+ 1" para incluir a última linha.
    cell1 = sheet.cell(row=row, column=column_number).value
    cell2 = sheet.cell(row=row + 1, column=column_number).value

    if cell1 is not None and (cell2 is None or cell1 != cell2):
        # Adicionar os valores às listas quando a próxima célula for nula ou diferente
        id.append(sheet.cell(row=row, column=1).value)
        tipo.append(sheet.cell(row=row, column=2).value)
        data.append(sheet.cell(row=row, column=3).value)


print("\n",id)
print("\n",tipo)
print("\n",data)

result_list = []


for row in range(1, sheet.max_row + 1):  # Note o "+ 1" para incluir a última linha.
    cell1 = sheet.cell(row=row, column=column_number).value
    cellD = sheet.cell(row=row, column=4).value  # Valor da coluna 'D'
    cellE = sheet.cell(row=row, column=5).value  # Valor da coluna 'E'
    cellF = sheet.cell(row=row, column=6).value  # Valor da coluna 'F'

    if cell1 is not None and (cellD is not None or cellE is not None or cellF is not None):
        # Adicionar os valores à lista id no formato desejado
        id.append([cell1, [cellD, [cellE, cellF]]])

print(id, end="\n")




'''
for item_id in id:
    for row in range(min_col=5, max_col=5,min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value is not None:
                        
                        
                        # Suponha que 'item_id' seja o valor da coluna A (por exemplo, 'PM-123')
                        col_D_value = sheet.cell(row=row, column=4).value  # Valor da coluna D
                        col_E_value = sheet.cell(row=row, column=5).value  # Valor da coluna E
                        col_F_value = sheet.cell(row=row, column=6).value  # Valor da coluna F
                        # Crie a estrutura de lista aninhada e adicione-a a 'result_list'
'''


# Agora, 'result_list' contém a estrutura desejada com os valores das colunas 'D', 'E' e 'F' para cada item em 'id'.

'''
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
             for cell in row:
                print(print(cell.value, end=' '))


  nested_list = [item_id, [col_D_value, col_E_value, col_F_value]]
    result_list.append(nested_list)
'''

print(f'Valores lidos em {sheet}:')

