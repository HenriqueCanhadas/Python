import openpyxl
import pandas as pd

workbook = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

n=70

for sheet_name in workbook.sheetnames:

    workbook = openpyxl.load_workbook('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    tabela = pd.read_excel('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx', sheet_name=sheet_name)

    sheet = workbook[sheet_name]

    identificacao  =sheet['C30'].value
    data =sheet['C6'].value
    horaensaio =sheet['H30'].value
    horaamostragem = sheet['K30'].value
    condutividade = tabela.iat[39, 4]#E41
    oxirreducao = tabela.iat[39, 6]#G41
    oxigenio = tabela.iat[39, 8]#I41
    ph = tabela.iat[39, 11]#L41
    temperatura = tabela.iat[39, 13]#N41
    turbidez = tabela.iat[39, 15]#P41
    condicoes =sheet['P30'].value


    condutividade = float(condutividade)
    oxirreducao = float(oxirreducao)
    oxigenio = float(oxigenio)
    ph = float(oxigenio)
    temperatura = float(temperatura)
    turbidez = float(turbidez)

    sheet = workbook['FINAL']

    sheet[f'A{n}']=identificacao
    sheet[f'B{n}']=data
    sheet[f'C{n}']=horaensaio
    sheet[f'D{n}']=horaamostragem
    sheet[f'E{n}']=condutividade
    sheet[f'F{n}']=oxirreducao
    sheet[f'G{n}']=oxigenio
    sheet[f'H{n}']=ph
    sheet[f'I{n}']=temperatura
    sheet[f'J{n}']=turbidez
    sheet[f'K{n}']=condicoes

    n+=1

    print(f'N esta em:{n}')

    workbook.save('Relatorio\REVISÃO - Boletim de Amostragem 7918.xlsx')

    print(identificacao)
    print(data)
    print(horaensaio)
    print(horaamostragem)
    print(condutividade)
    print(oxirreducao)
    print(oxigenio)
    print(ph)
    print(temperatura)
    print(turbidez)
    print(condicoes)





