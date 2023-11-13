import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Crie um novo arquivo Excel
workbook = openpyxl.Workbook()

# Selecione a planilha ativa
sheet = workbook.active

planilhap = 'P'

sheet['B1'] = 'Resultados das análises químicas de metais dissolvidos nas amostras de água subterrânea (µg/L) - Band'

sheet.merge_cells(f'B1:{planilhap}1')

sheet['B2'] = 'Parâmetros analisados'

sheet.merge_cells('B2:B6')

sheet['C2'] = 'Cas number'

sheet.merge_cells('C2:C6')

sheet['D2'] = 'Padrões de Referência  (µg/L)'

sheet.merge_cells('D2:E3')

sheet['D4'] = 'VI (1)'

sheet.merge_cells('D4:D6')

sheet['E4'] = 'RSL (3)'

sheet.merge_cells('E4:E6')

sheet['F2'] = 'Identificação da amostra/Data e hora de amostragem'

sheet.merge_cells(f'F2:{planilhap}3')

sheet['B7'] = 'Antimônio dissolvido '
sheet['B8'] = 'Arsênio dissolvido'
sheet['B9'] = 'Boro dissolvido '
sheet['B10'] = 'Bário dissolvido'
sheet['B11'] = 'Chumbo dissolvido'
sheet['B12'] = 'Cobalto dissolvido '
sheet['B13'] = 'Cobre dissolvido'
sheet['B14'] = 'Cromo dissolvido'
sheet['B15'] = 'Cádmio dissolvido'
sheet['B16'] = 'Mercúrio dissolvido'
sheet['B17'] = 'Molibdênio dissolvido '
sheet['B18'] = 'Níquel dissolvido'
sheet['B19'] = 'Prata dissolvida'
sheet['B20'] = 'Selênio dissolvido'
sheet['B21'] = 'Zinco dissolvido'

sheet['C7'] = '7440-36-0'
sheet['C8'] = '7440-38-2'
sheet['C9'] = '7440-42-8'
sheet['C10'] = '7440-39-3'
sheet['C11'] = '7439-92-1'
sheet['C12'] = '7440-48-4'
sheet['C13'] = '7440-50-8'
sheet['C14'] = '7440-47-3'
sheet['C15'] = '7440-43-9'
sheet['C16'] = '7439-97-6'
sheet['C17'] = '7439-98-7'
sheet['C18'] = '7440-02-0'
sheet['C19'] = '7440-22-4'
sheet['C20'] = '7782-49-2'
sheet['C21'] = '7440-66-6'

sheet['D7'] = '0,006'
sheet['D8'] = '0,01'
sheet['D9'] = '2,4'
sheet['D10'] = '0,7'
sheet['D11'] = '0,01'
sheet['D12'] = '0,07'
sheet['D13'] = '2'
sheet['D14'] = '0,05'
sheet['D15'] = '0,003'
sheet['D16'] = '0,001'
sheet['D17'] = '0,03'
sheet['D18'] = '0,07'
sheet['D19'] = '0,05'
sheet['D20'] = '0,04'
sheet['D21'] = '1,8'

sheet['E7'] = 'n.a'
sheet['E8'] = 'n.a'
sheet['E9'] = 'n.a'
sheet['E10'] = 'n.a'
sheet['E11'] = 'n.a'
sheet['E12'] = 'n.a'
sheet['E13'] = 'n.a'
sheet['E14'] = 'n.a'
sheet['E15'] = 'n.a'
sheet['E16'] = 'n.a'
sheet['E17'] = 'n.a'
sheet['E18'] = 'n.a'
sheet['E19'] = 'n.a'
sheet['E20'] = 'n.a'
sheet['E21'] = 'n.a'

Letra=["F","G","H","I","J","K","L","M","N","O","P"]

for i in Letra:
        sheet[f'{i}4'] = 'PM-10'
        sheet[f'{i}5'] = '11/07/2023 08:49:00'
        sheet.merge_cells(f'{i}5:{i}6')
        sheet[f'{i}7'] = '< 0,00500'
        sheet[f'{i}8'] = '< 0,00600'
        sheet[f'{i}9'] = '< 0,00600'
        sheet[f'{i}10'] = '< 0,00600'
        sheet[f'{i}11'] = '< 0,00600'
        sheet[f'{i}12'] = '< 0,00600'
        sheet[f'{i}13'] = '< 0,00600'
        sheet[f'{i}14'] = '< 0,00600'
        sheet[f'{i}15'] = '< 0,00020'
        sheet[f'{i}16'] = '< 0,00010'
        sheet[f'{i}17'] = '< 0,00600'
        sheet[f'{i}18'] = '< 0,00600'
        sheet[f'{i}19'] = '< 0,00500'
        sheet[f'{i}20'] = '< 0,00600'
        sheet[f'{i}21'] = '0,0101'   

workbook.save('TesteFormatacao/exemplo_formataçãoServmar.xlsx')