import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Crie um novo arquivo Excel
workbook = openpyxl.Workbook()

# Selecione a planilha ativa
sheet = workbook.active

# Adicione dados à planilha
sheet['A1'] = 'ID Ponto'
sheet['A2'] = 'Data de amostragem'
sheet['A3'] = 'ID Amostra'
sheet['A4'] = 'Relatório de Ensaio'

# Mesclar células e estilizar cabeçalhos
header_cell = sheet['A1']
header_cell.font = Font(bold=True)

sheet.merge_cells('A1:E1')
sheet.merge_cells('A2:E2')
sheet.merge_cells('A3:E3')
sheet.merge_cells('A4:E4')

sheet['A5'] = 'Parâmetro'
sheet['B5'] = 'CAS'
sheet['C5'] = 'Unidade'
sheet['D5'] = 'Valor Orientador'
sheet['E5'] = 'Referência'

# Pinte as células de verde e adicione bordas
green_fill = PatternFill(start_color='566b52', end_color='566b52', fill_type='solid')

border_style = Side(border_style='thin', color='000000')
for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=5):
    for cell in row:
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        cell.font = Font(color='FFFFFF')  # Defina a cor do texto como branco

for row in sheet.iter_rows(min_row=5, max_row=5, min_col=1, max_col=5):
    for cell in row:
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        cell.font = Font(color='FFFFFF')  # Defina a cor do texto como branco

sheet['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Salve o arquivo
workbook.save('TesteFormatacao/exemplo_formatação.xlsx')