# Criar um novo DataFrame para armazenar os resultados filtrados
#resultado_final = pd.DataFrame()
#
# Iterar sobre as listas e adicionar os resultados filtrados ao novo DataFrame
# for pm in lista_pm:
#    for descricao_metodo in lista_descricao_metodo:
#        filtro = (data_frame['SAMPLENAME'] == pm) & (data_frame['Descrição Método'] == descricao_metodo)
#        resultado_filtrado = data_frame[filtro]
#
#        # Verificar se o resultado não está vazio antes de adicionar ao DataFrame final
#        if not resultado_filtrado.empty:
#            resultado_final = resultado_final.append(resultado_filtrado[["SAMPLENAME", "SAMPDATE", "Descrição Método", "Análise", "CASNUMBER", "Result", "UNITS"]], ignore_index=True)
#
# for nome_da_aba in lista_descricao_metodo:
##    nova_aba = resultado_final.create_sheet(title=nome_da_aba)
#
#resultado_final.to_excel('C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Resultado_Final.xlsx', index=False)

import pandas as pd
from openpyxl import Workbook

# Lê a tabela "Analise Ceimic.xlsx" e a tabela "banco de dados - CEIMIC.xlsx"
tabela_lab = pd.read_excel(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Analise Ceimic.xlsx')
tabela_bd = pd.read_excel(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/banco de dados - CEIMIC.xlsx')

# Pega todos os itens que tem em comum em relação a coluna "Análise", junta em uma planilha só e salva como "Resultado_Merge.xlsx"
tabela_merge = pd.merge(tabela_lab, tabela_bd, how='left', on="Análise")
tabela_merge.to_excel(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Resultado_Merge.xlsx', index=False)
data_frame = pd.read_excel(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Resultado_Merge.xlsx')

# Pega da coluna "SAMPLENAME" todos os pms e pega da coluna "Descrição Método" todos as descrições dos métodos (pega os valores unicos e joga na lista)
lista_pm = data_frame['SAMPLENAME'].unique()
lista_descricao_metodo = data_frame['Descrição Método'].unique()

# Cria um novo DataFrame para armazenar os resultados filtrados e cria um arquivo excel no openpyxl
resultado_final = pd.DataFrame()
workbook = Workbook()

# Iterar sobre as listas e adicionar os resultados filtrados ao novo DataFrame
for descricao_metodo in lista_descricao_metodo:  # para cada descrição do metodo na lista
    # filtra na coluna descrição do método do arquivo "Resultado_Merge.xlsx" cada item na lista descrição do metodo
    filtro = (data_frame['Descrição Método'] == descricao_metodo)
    resultado_filtrado = data_frame[filtro]

    # Verificar se o resultado não está vazio antes de adicionar ao DataFrame final
    if not resultado_filtrado.empty:
        # Adicionar os resultados filtrados ao DataFrame final
        resultado_final = resultado_final.append(resultado_filtrado[["SAMPLENAME", "SAMPDATE", "Descrição Método", "Análise", "CASNUMBER", "Result", "UNITS"]], ignore_index=True)
        # Criar uma nova aba no Excel com o nome do item da lista_descricao_metodo
        aba_descricao_metodo = workbook.create_sheet(title=descricao_metodo)

        # Adiciona cada linha do descrição metdo em uma nova aba
        for linha in resultado_filtrado.itertuples(index=False):
            aba_descricao_metodo.append(linha)

# Salvar o novo DataFrame em uma nova planilha do Excel
resultado_final.to_excel(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Resultado_Final.xlsx', index=False)

# Salvar o Workbook do openpyxl no arquivo Excel
workbook.save(
    'C:/Users/vitor.lucas/Desktop/teste_CEIMIC/Resultado_Final_Com_Abas.xlsx')
