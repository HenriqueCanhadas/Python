from openpyxl import Workbook
import pandas as pd
'''
excel_laboratorio=openpyxl.load_workbook('Ceimic/Analise Ceimic.xlsx')

planilha=excel_laboratorio['Analise Ceimic']

Teste = []

pm = []  # antigo id
tipo_de_amostra = []  # antigo tipo
data_hora = []  # antiga data

for linha in range(2, planilha.max_row + 1):
        celula_pm = planilha.cell(row=linha, column=5).value  # Antiga cell1
        celula_seguinte = planilha.cell(row=linha+1, column=5).value  # Antiga cell2

        if celula_pm is not None and (celula_seguinte is None or celula_pm != celula_seguinte):
            pm.append(planilha.cell(row=linha, column=5).value)
            tipo_de_amostra.append(planilha.cell(row=linha, column=6).value)
            data_hora.append(planilha.cell(row=linha, column=6).value)

print(pm)

analises={}
resultado={}

for linha in range(2, planilha.max_row + 1):
        # Valor da coluna 'A' IDENTIFICAÇÃO DA AMOSTRA
        celula_pm = planilha.cell(row=linha, column=5).value
        # Valor da coluna 'D' ANALISE
        celula_analise = planilha.cell(row=linha, column=14).value
        # Valor da coluna 'E' RESULTADO
        celula_resultado = planilha.cell(row=linha, column=15).value
        # Valor da coluna 'F' UNIDADE
        celula_unidade = planilha.cell(row=linha, column=16).value

        # Check if celula_pm is not None and at least one of the analysis-related cells is not None
        if celula_pm is not None and (celula_analise is not None or celula_resultado is not None or celula_unidade is not None):
            # If the key (celula_pm) is not in analises, initialize it as an empty list
            if celula_pm not in analises:
                analises[celula_pm] = []
                resultado[celula_pm] = []

            analises[celula_pm].append(celula_analise)
            resultado[celula_pm].append([celula_resultado,celula_unidade])

print(analises['BV-04'])

quantidade_itens_bv_03 = len(analises['BV-04'])
print(f"Quantidade de itens: {quantidade_itens_bv_03}")

excel_laboratorio.close()

print("teste")
'''
 
tabela_lab = pd.read_excel('Ceimic/Analise Ceimic.xlsx')
tabela_bd = pd.read_excel('Ceimic/banco de dados - ceimic.xlsx')
 
tabela_merge = pd.merge(tabela_lab, tabela_bd,how='left', on="Análise") 
tabela_merge.to_excel('Ceimic/Resultado_Merge.xlsx', index=False)
data_frame = pd.read_excel('Ceimic/Resultado_Merge.xlsx')
 
lista_pm = data_frame['SAMPLENAME'].unique()
lista_descricao_metodo = data_frame['Descrição Método'].unique()
 
# Criar um novo DataFrame para armazenar os resultados filtrados
resultado_final = pd.DataFrame()
workbook = Workbook()
 
# Iterar sobre as listas e adicionar os resultados filtrados ao novo DataFrame
for descricao_metodo in lista_descricao_metodo:  # para cada descrição do metodo na lista
    # filtra na coluna descrição do método do arquivo "Resultado_Merge.xlsx" cada item na lista descrição do metodo
    filtro = (data_frame['Descrição Método'] == descricao_metodo)
    resultado_filtrado = data_frame[filtro]
    
    # Verificar se o resultado não está vazio antes de adicionar ao DataFrame final
    if not resultado_filtrado.empty:
        # Adicionar os resultados filtrados ao DataFrame final
        resultado_final = pd.concat([resultado_final, resultado_filtrado[["SAMPLENAME", "SAMPDATE", "Descrição Método", "Análise", "CASNUMBER", "Result", "UNITS"]]], ignore_index=True)
        # Criar uma nova aba no Excel com o nome do item da lista_descricao_metodo
        aba_descricao_metodo = workbook.create_sheet(title=descricao_metodo)

        # Adiciona cada linha do descrição metdo em uma nova aba
        for linha in resultado_filtrado.itertuples(index=False):
            aba_descricao_metodo.append(linha)

resultado_final.to_excel('Ceimic/Resultado_Final.xlsx', index=False)

workbook.save('Resultado_Final_Com_Abas.xlsx')